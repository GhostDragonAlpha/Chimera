"""ingest_gait_osf.py -- turn 246 measured adults into a table this story can read.

WHY THIS EXISTS. `theHuman` walked on `swing * sin(phase)` with `swing = 0.42 rad` -- a shape nobody
measured and an amplitude nobody sourced, and every membrane below inherited it: the vault, the
centre-of-pressure travel, the stride. A sine is not a hip. It was the one number in the whole gait
that was neither derived from a law nor read off a measurement, and it was doing most of the work.

THE SOURCE. Van Criekinge T, Saeys W, Truijen S et al., a normative 3D gait dataset of 246 healthy
adults aged 18-91, walking at three self-selected speeds (slow / comfortable / fast) over a CAREN
instrumented treadmill. Full-body kinematics, joint moments and powers, ground reaction force, and
spatiotemporal parameters, averaged over every valid stride and grouped by SEX and by DECADE of age.
CC BY 4.0. OSF doi 10.17605/OSF.IO/T72CW.

WHAT IT PRODUCES. `story/data/gait_normative.json`, a compact table of

    curve[speed][group][parameter] -> 100 samples of the gait cycle, mean and SD

with 1% of the cycle being heel strike of that leg and 100% the next heel strike of the same leg.
Twelve groups -- six age decades x two sexes -- so AGE AND SEX BECOME DIALS. Turn one and the hip
curve changes shape, the cadence changes, the step width changes, and everything the hierarchy
derives from those changes with it. That is the difference between having data and having a source:
a source has a range, and a range is a dial.

WHAT IS DELIBERATELY LEFT OUT. Joint powers, transverse-plane rotations, the left leg (identical to
the right within the reported asymmetry of ~3%), and the per-subject files. They are one download
away from the same script if a membrane ever needs them, and carrying 80 MB of spreadsheet for
curves nothing reads would be hoarding rather than sourcing.

RUN:  python tools/ingest_gait_osf.py            (downloads to research_references/human/gait_osf/)
      python tools/ingest_gait_osf.py --check    (re-derives and diffs, without writing)
"""
from __future__ import annotations

import argparse
import json
import sys
import urllib.request
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "research_references" / "human" / "gait_osf"
OUT = ROOT / "story" / "data" / "gait_normative.json"

DOI = "10.17605/OSF.IO/T72CW"
CITE = ("Van Criekinge T, Saeys W, Truijen S, et al. A full-body motion capture gait dataset of "
        "246 healthy adults aged 18-91 (2023). OSF, CC BY 4.0, doi " + DOI + ". "
        "Three self-selected speeds on a CAREN instrumented treadmill; group means and SD by sex "
        "and age decade, averaged over every valid stride.")

# The OSF file ids for the group-average workbooks, one per walking speed. These are the small
# (~5 MB) files: group means, not per-subject. The 20 MB per-subject Overview files are not needed
# for a normative curve and are not fetched.
FILES = {
    "comf": ("h8myj", "AgeGenderGroup_comf.xlsx"),
    "fast": ("6qx84", "AgeGenderGroup_fast.xlsx"),
    "slow": ("2s7bx", "AgeGenderGroup_slow.xlsx"),
    "demo": ("a7c5h", "Demo_PhysEx.xlsx"),
}

# The twelve groups, in the column order the workbooks use: for each age decade, men then women,
# each contributing a mean column and an SD column.
GROUPS = [f"{s}_{a}" for a in ("18-29", "30-39", "40-49", "50-59", "60-69", "70+")
          for s in ("m", "w")]

# WHAT EACH CURVE IS FOR, named by the membrane that needs it rather than by the sheet it lives in.
# A sheet name is the dataset's vocabulary; these are the story's.
#
# THE SIGN CONVENTION IS NOT TYPED HERE, AND THAT IS DELIBERATE. It is read out of each sheet's own
# A1 cell at ingest time. The dataset's description document and the spreadsheets DISAGREE about the
# ankle moment -- the document says "+ dorsal flexion", the sheet says "+ plantar flexion" -- and a
# transcribed convention is a comment claiming to be a measurement, which is the one failure this
# whole data layer exists to prevent. Whatever the file says travels with the numbers.
CURVES = {
    # the sagittal pose -- what a walker actually draws
    "hip_flex":     ("Rotation_RHipFlex_{s}",   "deg"),
    "knee_flex":    ("Rotation_RKneeFlex_{s}",  "deg"),
    "ankle_flex":   ("Rotation_RAnkleFlex_{s}", "deg"),
    # the frontal plane -- theBalance, which is the plane a sagittal model cannot see at all
    "hip_abad":     ("Rotation_RHipAbAd_{s}",   "deg"),
    "pelvic_obl":   ("Rotation_PelvicObl_{s}",  "deg"),
    "trunk_flex":   ("Rotation_TrunkFlex_{s}",  "deg"),
    # what the ground does back -- the double hump nobody has to fit.
    # UNIT ESTABLISHED BY ARITHMETIC, not by the document: the curve peaks at 1.10 and reads exactly
    # 0.00 through swing. The document calls the scalar peak table "newton", but 1.10 N is not a
    # person standing on a plate -- for this cohort's mean man it would be 888 N. The CURVE sheets
    # are normalised to body weight; the scalar sheets are not.
    "grf_vert":     ("GRF_R_vert_{s}",          "BW"),
    # what the ankle is asked for -- theAnkle claims 1.51 N.m/kg from three inherited numbers,
    # and this is the measurement that either confirms it or does not.
    # Same arithmetic: a peak of 1.51 is Nm/kg, not Nm (~112 Nm for this cohort).
    "ankle_moment": ("Moment_RAnkleFlex_{s}",   "Nm/kg"),
}

# The spatiotemporal table: one value per group, not a curve. These are the numbers a membrane
# checks itself against -- cadence, stride, step width, duty factor, double support.
GAIT_SHEET = "Gait_{s}"


def _fetch(speed: str) -> Path:
    fid, name = FILES[speed]
    RAW.mkdir(parents=True, exist_ok=True)
    dst = RAW / name
    if dst.exists() and dst.stat().st_size > 1_000_000:
        return dst
    url = f"https://osf.io/download/{fid}/"
    print(f"  fetching {name} from {url}")
    with urllib.request.urlopen(url, timeout=300) as r, open(dst, "wb") as f:
        f.write(r.read())
    print(f"  {name}: {dst.stat().st_size:,} bytes")
    return dst


def _curve(ws) -> dict:
    """Read one joint-angle sheet: 100 samples of the cycle x 12 groups x (mean, SD).

    LAYOUT, from the dataset's own description document: row 3 is the header, rows 4-103 are the
    cycle from 1% to 100%, column B is the percentage, and columns C onward alternate mean/SD for
    men then women within each age decade."""
    rows = list(ws.iter_rows(min_row=4, max_row=103, min_col=2, max_col=2 + 2 * len(GROUPS),
                             values_only=True))
    if len(rows) != 100:
        raise ValueError(f"expected 100 cycle samples, found {len(rows)}")
    out = {}
    for gi, g in enumerate(GROUPS):
        mean, sd = [], []
        for r in rows:
            m, s = r[1 + 2 * gi], r[2 + 2 * gi]
            mean.append(round(float(m), 2) if isinstance(m, (int, float)) else None)
            sd.append(round(float(s), 2) if isinstance(s, (int, float)) else None)
        out[g] = {"mean": mean, "sd": sd}
    return out


def _cohort(ws) -> dict:
    """WHO WAS MEASURED. Per-subject age, sex, mass, stature and leg length for all 246 adults,
    reduced to the statistics a membrane can actually use.

    THE LEG LENGTH COLUMN IS THE PRIZE HERE. `theHuman` carries leg length as a fraction of stature
    and that fraction was never sourced -- it was read off one model. This column is 246 people
    measured directly, so the fraction stops being an assertion and becomes a distribution with a
    spread, which is what lets a body be tall-legged or short-legged rather than average."""
    import statistics as st

    rows = [r for r in ws.iter_rows(min_row=5, max_col=10, values_only=True)
            if isinstance(r[1], str) and r[1].startswith("HAC")]
    num = lambda i, f=lambda r: True: [float(r[i]) for r in rows
                                       if isinstance(r[i], (int, float)) and f(r)]
    men, women = (lambda r: r[3] == 0), (lambda r: r[3] == 1)
    frac = [float(r[6]) / float(r[5]) for r in rows
            if isinstance(r[6], (int, float)) and isinstance(r[5], (int, float)) and r[5] > 1]
    frac_m = [float(r[6]) / float(r[5]) for r in rows if r[3] == 0
              and isinstance(r[6], (int, float)) and isinstance(r[5], (int, float)) and r[5] > 1]
    frac_w = [float(r[6]) / float(r[5]) for r in rows if r[3] == 1
              and isinstance(r[6], (int, float)) and isinstance(r[5], (int, float)) and r[5] > 1]
    ms = lambda v: [round(st.mean(v), 4), round(st.pstdev(v), 4), len(v)]
    return {
        "n": len(rows),
        "note": "mean, population SD, n -- for every entry below",
        "age_years": ms(num(2)), "men": sum(1 for r in rows if r[3] == 0),
        "women": sum(1 for r in rows if r[3] == 1),
        "mass_kg": ms(num(4)), "mass_kg_men": ms(num(4, men)), "mass_kg_women": ms(num(4, women)),
        "height_m": ms(num(5)), "height_m_men": ms(num(5, men)),
        "height_m_women": ms(num(5, women)),
        "leg_length_m": ms(num(6)),
        # the fraction, which is the thing a membrane scales by
        "leg_over_stature": ms(frac),
        "leg_over_stature_men": ms(frac_m),
        "leg_over_stature_women": ms(frac_w),
        "leg_length_definition": ("as measured by the study for its CAREN marker model; "
                                  "greater-trochanter/ASIS to malleolus, the length that swings"),
    }


def _spatiotemporal(ws) -> dict:
    """Read the Gait sheet: named scalar parameters, one row each, same 12-group column layout."""
    out = {}
    for r in ws.iter_rows(min_row=4, max_row=40, min_col=1, max_col=1 + 2 * len(GROUPS),
                          values_only=True):
        name = r[0]
        if not isinstance(name, str) or not name.strip():
            continue
        # names arrive as "L.Step.Width [m] " -- keep the unit, lose the whitespace
        key = name.strip()
        vals = {}
        for gi, g in enumerate(GROUPS):
            m, s = r[1 + 2 * gi], r[2 + 2 * gi]
            if isinstance(m, (int, float)):
                vals[g] = [round(float(m), 4),
                           round(float(s), 4) if isinstance(s, (int, float)) else None]
        if vals:
            out[key] = vals
    return out


def build() -> dict:
    import openpyxl

    data = {
        "source": CITE,
        "doi": DOI,
        "licence": "CC BY 4.0",
        "n_subjects": 246,
        "age_range_years": [18, 91],
        "groups": GROUPS,
        "cycle_samples": 100,
        "cycle_note": ("sample 1 is heel strike of the right leg, sample 100 is the next heel "
                       "strike of the same leg; percentages, not seconds"),
        "leg": "right",
        "curve_units": {k: v[1] for k, v in CURVES.items()},
        "curve_sign": {},   # filled from each sheet's own A1 label, never transcribed
        "curve_sheet": {k: v[0] for k, v in CURVES.items()},
        "speeds": {},
    }

    demo = _fetch("demo")
    wb = openpyxl.load_workbook(demo, read_only=True, data_only=True)
    data["cohort"] = _cohort(wb[wb.sheetnames[0]])
    wb.close()
    c = data["cohort"]
    print(f"  cohort: {c['n']} adults, {c['men']}M/{c['women']}W, age {c['age_years'][0]:.1f}"
          f"+-{c['age_years'][1]:.1f}, leg/stature {c['leg_over_stature'][0]:.4f}"
          f"+-{c['leg_over_stature'][1]:.4f}")

    for speed in ("slow", "comf", "fast"):
        path = _fetch(speed)
        print(f"  reading {path.name}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        block = {"spatiotemporal": _spatiotemporal(wb[GAIT_SHEET.format(s=speed)]), "curves": {}}
        for key, (sheet, _u) in CURVES.items():
            ws = wb[sheet.format(s=speed)]
            block["curves"][key] = _curve(ws)
            # the sheet's own sign convention, read rather than retyped
            a1 = ws.cell(row=1, column=1).value
            if isinstance(a1, str) and a1.strip():
                data["curve_sign"].setdefault(key, " ".join(a1.split()))
        wb.close()
        data["speeds"][speed] = block
        n = len(block["curves"]) * len(GROUPS) * 100
        print(f"    {speed}: {len(block['spatiotemporal'])} parameters, {n:,} curve samples")

    missing = [k for k in CURVES if k not in data["curve_sign"]]
    if missing:
        # An unlabelled sheet is not fatal, but it must be visible: a curve whose sign nobody stated
        # can be drawn upside down and still look like a walk.
        print(f"  NOTE: no sign label in the source for {', '.join(missing)}")
        for k in missing:
            data["curve_sign"][k] = "UNLABELLED IN SOURCE"
    return data


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--check", action="store_true",
                    help="re-derive and compare against the committed table, writing nothing")
    a = ap.parse_args()

    print(f"ingesting the OSF normative gait dataset (doi {DOI})")
    data = build()
    blob = json.dumps(data, separators=(",", ":"))

    if a.check:
        if not OUT.exists():
            print("  no committed table to check against")
            return 1
        same = OUT.read_text(encoding="utf8") == blob
        print(f"  {'MATCHES' if same else 'DIFFERS FROM'} {OUT.relative_to(ROOT)}")
        return 0 if same else 2

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(blob, encoding="utf8")
    print(f"  wrote {OUT.relative_to(ROOT)}  ({len(blob):,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
